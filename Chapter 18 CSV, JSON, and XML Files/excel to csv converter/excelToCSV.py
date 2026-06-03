#Chapter 18 of Automate the Boring Stuff by Al Sweigart
#https://automatetheboringstuff.com/3e/chapter18.html

import openpyxl, os, csv

folder = 'excelSpreadsheets'
csv_folder = 'csvOutput'

os.makedirs(csv_folder, exist_ok=True)

for excel_file in os.listdir(folder):
    # Skip non-xlsx files, load the workbook object.
    if not excel_file.endswith('.xlsx'):
        continue
    
    wb = openpyxl.load_workbook(os.path.join(folder, excel_file), data_only=True)

    for sheet_name in wb.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = wb[sheet_name]

        #Split the pathname path into a pair (root, ext) such that root + ext == path, and the extension, ext, is empty or begins with a period and contains at most one period.
        base_name = os.path.splitext(excel_file)[0]
        new_file_name = f'{base_name}_{sheet_name}.csv'

        # Create the CSV filename from the Excel filename and sheet title.
        csv_file = open(os.path.join(csv_folder, new_file_name), 'w', newline='')
        
        # Create the csv.writer object for this CSV file.
        output_writer = csv.writer(csv_file)
        
        # Loop through every row in the sheet.
        for row_num in range(1, sheet.max_row + 1):
            row_data = []    # Append each cell to this list.
            # Loop through each cell in the row.
            for col_num in range(1, sheet.max_column + 1):
                # Append each cell's data to row_data
                cell_data = sheet.cell(row=row_num, column=col_num).value
                row_data.append(cell_data)
           
            # Write the row_data list to the CSV file.
            output_writer.writerow(row_data)
        
        csv_file.close()
