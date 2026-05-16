import openpyxl, os
from pathlib import Path
import pprint

def get_home_folder_size():
    filenames_and_sizes = []

    # Loop over everything in the home folder:
    for filename in os.listdir(Path.home()):
        absolute_file_path = Path.home() / filename

        # Skip folders/directories:
        if absolute_file_path.is_dir():
            continue

        # Get file size:
        try:
            file_size = absolute_file_path.stat().st_size
        except Exception:
            # Skip files with permissions errors:
            continue

        # Record filename and size:
        filenames_and_sizes.append((filename, file_size))

    return filenames_and_sizes

# Uncomment to print the hundred largest filenames and sizes:
#pprint.pprint(get_home_folder_size())


def make_excel_report(filenames_and_sizes):
    wb = openpyxl.Workbook()
    sheet = wb.active
    for index, file in enumerate(filenames_and_sizes, start=1):
        sheet.cell(row = index, column = 1).value = file[0]
        sheet.cell(row = index, column = 2).value = file[1]
    
    wb.save('homeFilesReport.xlsx')


make_excel_report(get_home_folder_size())