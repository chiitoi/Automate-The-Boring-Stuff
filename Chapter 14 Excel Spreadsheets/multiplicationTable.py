#Chapter 14 of Automate the Boring Stuff by Al Sweigart
#https://automatetheboringstuff.com/3e/chapter14.html

import sys, openpyxl
from openpyxl.styles import Font

if len(sys.argv) > 1:
    if not sys.argv[1].isdigit():
        print('Please pass a number to the program.')
        sys.exit()
    max_number = int(sys.argv[1])
else:
    max_number = 6

wb = openpyxl.Workbook()
sheet = wb.active

bold_font = Font(bold=True)

for i in range(1, max_number + 1):
    #Column A
    cell_a = sheet.cell(row=i + 1, column=1)
    cell_a.value = i
    cell_a.font = bold_font

    #Row 1
    cell_row = sheet.cell(row=1, column=i + 1)
    cell_row.value = i
    cell_row.font = bold_font

for row_number in range(1, max_number+1):
    for column_number in range(1, max_number+1):
        cell = sheet.cell(row= row_number + 1, column= column_number + 1)
        cell.value = row_number*column_number

wb.save('multiplication.xlsx')